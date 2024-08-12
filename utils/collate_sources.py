import streamlit as st
from psycopg2.extensions import connection
from db.db import connect_to_db
from db.queries import get_db_domains
from psycopg2.extensions import connection as Psycopg2Connection
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
import openpyxl
from fuzzywuzzy import fuzz, process
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import regex as re
from utils.excel import read_worksheets

def normalise_domains(df: pd.DataFrame) -> pd.DataFrame:
    '''
        Normalise domains removes hhtps://, www and / from url
        
        Args
            dataframe: a dataframe of the worksheet
        Return
            dataframe
    '''

    extract_domain = lambda url: re.sub(r"^(https?://)?(www\.)?|(\/)+$", "", url)
    if "domain" in df.columns:
        df["domain"] = df["domain"].apply(
            lambda x: extract_domain(x) if x is not None else None
        )
        return df
    return df


def fuzzy_matching(value_string: str, match_against: pd.Series, threshold=90) -> str:
    '''We fuzzy match names of domains in each worksheet against names in other worksheets and return the best match

        Args
            value_string: the name of the domain
            match_against: name in other sheet to match against
            threshodl: default fuzzy matching threshold is 90
        Returns
            Returns the best matched name as string
    '''
    match, score, i = process.extractOne(value_string, match_against, scorer=fuzz.ratio)
    return match if score >= threshold else value_string


def create_match_col(df: pd.DataFrame, match_against: pd.DataFrame) -> pd.Series | None:
    '''
    We create a column for each worksheet with best match name and call fuzzy matching function
    Args
        df: dataframe to match names for
        match_against: the dataframe to match names against
    Return
    Returns a column containing best matches

    '''
    try:
        name_series = df["name"].apply(lambda x: fuzzy_matching(x, match_against["name"]))
        #domain_series = df["domain"].apply(lambda x: fuzzy_matching(x, match_against["domain"]))
        return name_series #, domain_series
    except Exception as e:
        st.toast(f"Ensure there are no media sources with a missing name: {e}")
        st.stop()

def match_names_domains(worksheets: dict) -> dict | None:
    '''Calls best matches by calling the create match col which calls fuzzy matching'''
    for key, df in worksheets.items():
        # if key != 'comply':
        name_series: pd.Series = create_match_col(df, worksheets["comply"])
        df["match_name"] = name_series
        worksheets[key] = df

    return worksheets

def collate_sheets(worksheets: dict) -> pd.DataFrame | None:
    '''Collate sheets concats sheets and groups them by match_name'''
    try:
        # Types in Excel require exact same so circulations that are flaots and str break programme. Cast everything as str
        concat_data = pd.concat(worksheets.values()).astype(str)
        merged_df = concat_data.groupby("match_name").agg(tuple).applymap(list).reset_index()
        return merged_df
    except Exception as e:
        st.toast(f"⚠️ Failed to collate {e}")

def tidy_taxonomy_col(lst):
    #Remove NaN values (floats)
    lst = [x for x in lst if not (isinstance(x, float) and np.isnan(x))]
    
    #Remove None values
    lst = [x for x in lst if x not in ["None",None, ""]]

    #Remove duplicates and convert to comma-separated string
    unique = set(lst)
    return ", ".join(str(item) for item in unique)

def collate_sources(workbook: BytesIO) -> pd.DataFrame | None:
    '''
        Collate sources is the main function for collation. It reads all worksheets except main and collates them into main sheet.
        1. The collate sources reads worksheets into dataframes.
        2. Collator then normalises domains, performs fuzzy matching. 
        3. Finally, dataframes are concatenated into one df and grouped by best_match names. 
        4. Result is a dataframe representation of our source of truth.

        Args
            workbook: the BytesIO workbook in streamlit store
        Returns:
            dataframe representation of main sheet.
    '''
    workbook.seek(0)
    workbook: Workbook = openpyxl.load_workbook(workbook)
    sheet_names: list[str] = workbook.sheetnames
    worksheet_list: list[Worksheet] = workbook.worksheets
    worksheets: dict = read_worksheets(sheet_names, worksheet_list)

    # Clean the domains
    for df in worksheets.values():
        df = normalise_domains(df)
    
    worksheets: dict = match_names_domains(worksheets)
    df: pd.DataFrame | None = collate_sheets(worksheets) 

    if all(col in df.columns for col in ['taxonomy']):
        df['taxonomy'] = df['taxonomy'].apply(lambda x: tidy_taxonomy_col(x))

    st.toast("✔️ Successfully collated sources")
    return df
