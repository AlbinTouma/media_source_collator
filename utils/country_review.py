import streamlit as st
from psycopg2.extensions import connection
from db.db import connect_to_db, create_temporary_table, query_for_mentions
from db.queries import get_db_domains
from psycopg2.extensions import connection as Psycopg2Connection
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import openpyxl
from fuzzywuzzy import fuzz, process
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import regex as re


def db_get_country_domains(country: str) -> list[tuple] | Exception:
    conn: Psycopg2Connection | Exception = connect_to_db(db_name="Source Metadata")
    if isinstance(conn, Psycopg2Connection):
        try:
            cur = conn.cursor()
            cur.execute(get_db_domains, (country,))
            sql_response = cur.fetchall()
            colnames = [desc[0] for desc in cur.description]
            return pd.DataFrame(sql_response, columns=colnames)
        except Exception as e:
            st.warning(str(e))
            conn.rollback()
        finally:
            conn.close()


def init_workbook(file=None) -> BytesIO:
    if file:
        return BytesIO(file.read())
    else:
        wb = openpyxl.Workbook()

        # Set the default Sheet as main so we can overwrite comply and avoid empty sheet errors.
        for w in wb.worksheets:
            w.title = "comply"

        stream = BytesIO()
        wb.save(stream)

        stream.seek(0)
        return stream


def excel_writer(
    workbook_stream: BytesIO, sheet_name: str, df: pd.DataFrame
) -> BytesIO:

    #Excel doesn't accept NaN so replacing with None
    df = df.where(pd.notnull(df), "")

    workbook_stream.seek(0)
    workbook = openpyxl.load_workbook(workbook_stream)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)

    # Write columns
    for col_idx, col_name in enumerate(df.columns, 1):
        sheet.cell(row=1, column=col_idx, value=col_name)

    # Write rows
    '''Dealing with ValueError cannot convert[nan] to Excel indicates that there is an issue with converting a list containing nan. Excel doesn't support NaN so we convert [nan, nan, x] to str.'''


    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, list):
                value = ', '.join(map(str, value))
            if value is None:
                value = ''

            sheet.cell(row=r_idx, column=c_idx, value=value)

    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    return output_stream


def worksheet_to_dataframe(worksheet: Worksheet) -> pd.DataFrame:
    """Worksheet.values produces a cell values by row as tuple"""

    data: Iterator[Tuple] = worksheet.values
    data_list: list[Iterator] = list(data)

    # Check if the worksheet is empty and return empty df
    if len(data_list) == 0:
        return pd.DataFrame()

    columns = data_list[0]
   
    # Remove rows that are all empty to avoid empty values in df
    filtered_rows = [row for row in data_list[1:] if any(cell is not None and cell != '' for cell in row)]  
    df = pd.DataFrame(filtered_rows, columns=columns)
    return df


def read_worksheets(
    sheet_names, work_sheets: list[Worksheet]
) -> dict[str, pd.DataFrame]:

    data = {}
    for name, worksheet in zip(sheet_names, work_sheets):
        if name == "main":
            continue
        data[name] = worksheet_to_dataframe(worksheet)

    return data


def normalise_domains(df: pd.DataFrame) -> pd.DataFrame:
    extract_domain = lambda url: re.sub(r"^(https?://)?(www\.)?|(\/)+$", "", url)
    if "domain" in df.columns:
        df["domain"] = df["domain"].apply(
            lambda x: extract_domain(x) if x is not None else None
        )
        return df
    return df


def fuzzy_matching(value_string: str, match_against: pd.Series, threshold=90) -> str:
    match, score, i = process.extractOne(value_string, match_against, scorer=fuzz.ratio)
    return match if score >= threshold else value_string


def create_match_col(df: pd.DataFrame, match_against: pd.DataFrame) -> pd.Series | None:
    try:
        name_series = df["name"].apply(lambda x: fuzzy_matching(x, match_against["name"]))
        #domain_series = df["domain"].apply(lambda x: fuzzy_matching(x, match_against["domain"]))
        return name_series #, domain_series
    except Exception as e:
        st.warning(f"Ensure there are no media sources with a missing name: {e}")
        st.stop()

def match_names_domains(worksheets: dict) -> dict | None:
    for key, df in worksheets.items():
        # if key != 'comply':
        name_series: pd.Series = create_match_col(df, worksheets["comply"])
        df["match_name"] = name_series
        worksheets[key] = df

    return worksheets

def collate_sheets(worksheets: dict) -> pd.DataFrame | None:
    try:
        # Types in Excel require exact same so circulations that are flaots and str break programme. Cast everything as str
        concat_data = pd.concat(worksheets.values()).astype(str)
        merged_df = concat_data.groupby("match_name").agg(tuple).applymap(list).reset_index()
        return merged_df
    except Exception as e:
        st.toast(f"⚠️ Failed to collate {e}")


def collate_sources(workbook: BytesIO) -> pd.DataFrame | None:
    workbook.seek(0)
    workbook: Workbook = openpyxl.load_workbook(workbook)
    sheet_names: list[str] = workbook.sheetnames
    worksheet_list: list[Worksheet] = workbook.worksheets
    worksheets: dict = read_worksheets(sheet_names, worksheet_list)

    # Clean the domains
    for df in worksheets.values():
        df = normalise_domains(df)
    
    worksheets: dict = match_names_domains(worksheets)
    df: pd.DataFrame | None = collate_sheets(worksheets) 
    st.toast("✔️ Successfully collated sources")
    return df
