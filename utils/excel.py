import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import regex as re


def init_workbook(file=None) -> BytesIO:
    '''Init workbook creates and saves workbook for streamlit session.
    If user uploads workbook it is read and saved as BytesIO.
    '''

    if file:
        return BytesIO(file.read())
    else:
        wb = openpyxl.Workbook()

        # Set the default Sheet as main so we can overwrite comply and avoid empty sheet errors.
        for w in wb.worksheets:
            w.title = "comply"

        stream = BytesIO()
        wb.save(stream)

        stream.seek(0)
        return stream


def excel_writer(
    workbook_stream: BytesIO, sheet_name: str, df: pd.DataFrame
) -> BytesIO:
    '''
    Writes data to sheet in our excel BytesIO object.

    Args: 
        workbook_stream: workbook as BytesIO
        sheet_name: name of the sheet to write data to ie comply
        df: dataframe ie comply db data is saved to df and passed to excel writer

    Returns:
        BytesIO: a BytesIO rep of the excel with data written to sheet.

    '''

    #Excel doesn't accept NaN so replacing with None
    df = df.where(pd.notnull(df), "")

    workbook_stream.seek(0)
    workbook = openpyxl.load_workbook(workbook_stream)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    sheet = workbook.create_sheet(title=sheet_name)

    # Write columns
    for col_idx, col_name in enumerate(df.columns, 1):
        sheet.cell(row=1, column=col_idx, value=col_name)

    # Write rows
    '''Dealing with ValueError cannot convert[nan] to Excel indicates that there is an issue with converting a list containing nan. Excel doesn't support NaN so we convert [nan, nan, x] to str.'''


    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            if isinstance(value, list):
                value = ', '.join(map(str, value))
            if value is None:
                value = ''

            sheet.cell(row=r_idx, column=c_idx, value=value)

    output_stream = BytesIO()
    workbook.save(output_stream)
    output_stream.seek(0)
    return output_stream


def worksheet_to_dataframe(worksheet: Worksheet) -> pd.DataFrame:
    """
    Reads worksheet into dataframe. Part of the read_worksheet function.
    This function reads Worksheet.values (row in worksheet is a tuple) and removes empty rows
    before writing them to a dataframe.

    Args:
        worksheet: Worksheet object
    Returns
        dataframe of the worksheet.

    """

    data: Iterator[Tuple] = worksheet.values
    data_list: list[Iterator] = list(data)

    # Check if the worksheet is empty and return empty df
    if len(data_list) == 0:
        return pd.DataFrame()

    columns = data_list[0]
   
    # Remove rows that are all empty to avoid empty values in df
    filtered_rows = [row for row in data_list[1:] if any(cell is not None and cell != '' for cell in row)]  
    df = pd.DataFrame(filtered_rows, columns=columns)
    return df


def read_worksheets(
    sheet_names: list[str], work_sheets: list[Worksheet]
) -> dict[str, pd.DataFrame]:
    '''
        Reads worksheets in a workbook (except worksheet main). Used in collator.

        Args
            sheet_names: list of names of worksheets
            work_sheets: list of Worksheet objects
        Returns
            dictionary of worksheet names and dataframe representations of the worksheet.

        '''

    data = {}
    for name, worksheet in zip(sheet_names, work_sheets):
        if name == "main":
            continue
        data[name] = worksheet_to_dataframe(worksheet)

    return data

